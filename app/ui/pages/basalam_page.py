from __future__ import annotations

import json
import logging
import math
import time

import requests
from PySide6.QtCore import QObject, Qt, QThread, QTime, Signal, Slot
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTimeEdit,
    QVBoxLayout,
    QWidget,
)

from app.core.config import AppConfig
from app.services.basalam_service import list_vendor_orders
from app.services.basalam_store import BasalamIdStore
from app.utils import dialogs
from app.utils.dates import jalali_month_days, jalali_to_gregorian, jalali_today
from app.utils.numeric import format_amount, is_price_column

PERSIAN_MONTHS = [
    "Farvardin",
    "Ordibehesht",
    "Khordad",
    "Tir",
    "Mordad",
    "Shahrivar",
    "Mehr",
    "Aban",
    "Azar",
    "Dey",
    "Bahman",
    "Esfand",
]

TARGET_STATUS_FA = "رضایت مشتری"
TAB_COMPLETED = "COMPLECTED"
TABS_TO_FETCH = ["SHIPPED", "PENDING", TAB_COMPLETED]
PAGE_LIMIT = 30
REQUEST_SLEEP_SECONDS = 5


def extract_records(payload) -> list[dict]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        data = payload.get("data")
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return [data]
        for key in ("items", "results", "parcels", "orders"):
            value = payload.get(key)
            if isinstance(value, list):
                return value
        for key in ("items", "results", "parcels", "orders"):
            value = data.get(key) if isinstance(data, dict) else None
            if isinstance(value, list):
                return value
        return [payload]
    return []


def status_matches(record: dict, target: str) -> bool:
    if not isinstance(record, dict):
        return False
    for key, value in record.items():
        key_lower = str(key).lower()
        if (
            "status" in key_lower
            or "state" in key_lower
            or "وضعیت" in key_lower
        ):
            if value_matches_status(value, target):
                return True
    return False


def value_matches_status(value, target: str) -> bool:
    if isinstance(value, str):
        return value.strip() == target
    if isinstance(value, dict):
        for key in ("title", "name", "label", "value", "fa", "fa_IR"):
            inner = value.get(key)
            if isinstance(inner, str) and inner.strip() == target:
                return True
        for key, inner in value.items():
            if "status" in str(key).lower() and isinstance(inner, str):
                if inner.strip() == target:
                    return True
    return False


class BasalamWorker(QObject):
    progress = Signal(int)
    finished = Signal(list, int, int)
    error = Signal(str)

    def __init__(
        self,
        vendor_id: str,
        start_paid_at: str,
        end_paid_at: str,
        access_token: str,
        tabs: list[str] | None = None,
        completed_status: str = TARGET_STATUS_FA,
        sleep_seconds: int = REQUEST_SLEEP_SECONDS,
    ) -> None:
        super().__init__()
        self.vendor_id = vendor_id
        self.start_paid_at = start_paid_at
        self.end_paid_at = end_paid_at
        self.access_token = access_token
        self.tabs = tabs or list(TABS_TO_FETCH)
        self.completed_status = completed_status
        self.sleep_seconds = sleep_seconds
        self._logger = logging.getLogger(self.__class__.__name__)
        self._id_store = BasalamIdStore()

    @Slot()
    def run(self) -> None:
        try:
            records, total_count, skipped_existing = self._fetch_all_records()
        except requests.HTTPError as exc:
            message = str(exc)
            if exc.response is not None and exc.response.text:
                message = f"{message}\n\n{exc.response.text}"
            self._logger.exception("Basalam worker HTTP error")
            self.error.emit(message)
            return
        except requests.RequestException as exc:
            self._logger.exception("Basalam worker request failed")
            self.error.emit(str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            self._logger.exception("Basalam worker crashed")
            self.error.emit(str(exc))
            return

        self.finished.emit(records, total_count, skipped_existing)

    def _fetch_all_records(self) -> tuple[list[dict], int, int]:
        all_records: list[dict] = []
        seen_ids: set[str] = set()
        fetched_total = 0
        skipped_existing = 0
        stored_ids: list[str] = []
        self._logger.info(
            "Basalam worker start vendor=%s start=%s end=%s tabs=%s",
            self.vendor_id,
            self.start_paid_at,
            self.end_paid_at,
            ",".join(self.tabs),
        )

        for tab in self.tabs:
            offset = 0
            while True:
                payload = list_vendor_orders(
                    vendor_id=self.vendor_id,
                    tab=tab,
                    start_paid_at=self.start_paid_at,
                    end_paid_at=self.end_paid_at,
                    limit=PAGE_LIMIT,
                    offset=offset,
                    access_token=self.access_token,
                )
                batch = extract_records(payload)
                raw_batch_len = len(batch)
                if not batch:
                    break

                fetched_total += raw_batch_len
                self.progress.emit(fetched_total)

                if tab == TAB_COMPLETED:
                    batch = [
                        item
                        for item in batch
                        if status_matches(item, self.completed_status)
                    ]

                unique_batch: list[dict] = []
                batch_ids: list[str] = []
                for item in batch:
                    if not isinstance(item, dict):
                        continue
                    item_id = str(item.get("id", "")).strip()
                    if item_id:
                        if item_id in seen_ids:
                            continue
                        seen_ids.add(item_id)
                        batch_ids.append(item_id)
                    unique_batch.append(item)

                existing_ids = (
                    self._id_store.fetch_existing_ids(batch_ids)
                    if batch_ids
                    else set()
                )
                for item in unique_batch:
                    item_id = str(item.get("id", "")).strip()
                    if item_id and item_id in existing_ids:
                        skipped_existing += 1
                        continue
                    all_records.append(item)
                    if item_id:
                        stored_ids.append(item_id)

                if raw_batch_len < PAGE_LIMIT:
                    break

                offset += PAGE_LIMIT
                time.sleep(self.sleep_seconds)

        if stored_ids:
            self._id_store.store_ids(stored_ids)

        self._logger.info(
            "Basalam worker finished total=%s skipped_existing=%s",
            len(all_records),
            skipped_existing,
        )
        return (
            all_records,
            len(all_records) + skipped_existing,
            skipped_existing,
        )


class JalaliDateTimePicker(QWidget):
    def __init__(
        self,
        parent: QWidget | None = None,
        default_time: QTime | None = None,
    ) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.year_combo = QComboBox()
        self.month_combo = QComboBox()
        self.day_combo = QComboBox()
        self.time_edit = QTimeEdit()
        self.time_edit.setDisplayFormat("HH:mm:ss")

        for year in range(1390, 1451):
            self.year_combo.addItem(str(year), year)
        for idx, name in enumerate(PERSIAN_MONTHS, start=1):
            self.month_combo.addItem(f"{idx:02d} {name}", idx)

        layout.addWidget(self.year_combo)
        layout.addWidget(self.month_combo)
        layout.addWidget(self.day_combo)
        layout.addWidget(self.time_edit)

        self.year_combo.currentIndexChanged.connect(self._refresh_days)
        self.month_combo.currentIndexChanged.connect(self._refresh_days)

        jy, jm, jd = jalali_today()
        self._set_date(jy, jm, jd)
        if default_time is not None:
            self.time_edit.setTime(default_time)

    def _set_date(self, jy: int, jm: int, jd: int) -> None:
        self.year_combo.setCurrentText(str(jy))
        month_index = max(1, min(jm, 12)) - 1
        self.month_combo.setCurrentIndex(month_index)
        self._refresh_days()
        day_index = max(1, min(jd, self.day_combo.count())) - 1
        self.day_combo.setCurrentIndex(day_index)

    def _refresh_days(self) -> None:
        jy = int(self.year_combo.currentData())
        jm = int(self.month_combo.currentData())
        current_day = self.day_combo.currentData()
        max_day = jalali_month_days(jy, jm)

        self.day_combo.blockSignals(True)
        self.day_combo.clear()
        for day in range(1, max_day + 1):
            self.day_combo.addItem(f"{day:02d}", day)
        if isinstance(current_day, int) and 1 <= current_day <= max_day:
            self.day_combo.setCurrentIndex(current_day - 1)
        self.day_combo.blockSignals(False)

    def jalali_date(self) -> tuple[int, int, int]:
        jy = int(self.year_combo.currentData())
        jm = int(self.month_combo.currentData())
        jd = int(self.day_combo.currentData())
        return jy, jm, jd

    def to_gregorian_str(self) -> str:
        jy, jm, jd = self.jalali_date()
        gy, gm, gd = jalali_to_gregorian(jy, jm, jd)
        time_text = self.time_edit.time().toString("HH:mm:ss")
        return f"{gy:04d}-{gm:02d}-{gd:02d} {time_text}"


class BasalamPage(QWidget):
    def __init__(
        self, config: AppConfig, parent: QWidget | None = None
    ) -> None:
        super().__init__(parent)
        self.config = config
        self._dataframe = None
        self._logger = logging.getLogger(self.__class__.__name__)
        self._worker_thread: QThread | None = None
        self._worker: BasalamWorker | None = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        header = QHBoxLayout()
        title = QLabel("Basalam Orders")
        title.setStyleSheet("font-size: 16px; font-weight: 600;")
        header.addWidget(title)
        header.addStretch(1)

        self.fetch_button = QPushButton("Fetch")
        self.fetch_button.clicked.connect(self._fetch)
        header.addWidget(self.fetch_button)

        self.export_button = QPushButton("Export")
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(self._export)
        header.addWidget(self.export_button)
        layout.addLayout(header)

        progress_row = QHBoxLayout()
        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet("color: #6B7280;")
        self.progress_label.hide()
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)
        self.progress_bar.hide()
        progress_row.addWidget(self.progress_label)
        progress_row.addWidget(self.progress_bar, 1)
        layout.addLayout(progress_row)

        form_card = QFrame()
        form_card.setObjectName("Card")
        form_layout = QGridLayout(form_card)
        form_layout.setContentsMargins(16, 16, 16, 16)
        form_layout.setHorizontalSpacing(12)
        form_layout.setVerticalSpacing(12)

        start_paid_label = QLabel("Start paid at (Jalali)")
        self.start_paid_input = JalaliDateTimePicker(
            default_time=QTime(0, 0, 0)
        )
        form_layout.addWidget(start_paid_label, 0, 0)
        form_layout.addWidget(self.start_paid_input, 0, 1)

        end_paid_label = QLabel("End paid at (Jalali)")
        self.end_paid_input = JalaliDateTimePicker(
            default_time=QTime(23, 59, 59)
        )
        form_layout.addWidget(end_paid_label, 1, 0)
        form_layout.addWidget(self.end_paid_input, 1, 1)

        layout.addWidget(form_card)

        self.summary_label = QLabel("No data loaded.")
        self.summary_label.setStyleSheet("color: #6B7280;")
        layout.addWidget(self.summary_label)

        table_card = QFrame()
        table_card.setObjectName("Card")
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(16, 16, 16, 16)

        self.table = QTableWidget(0, 0)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSortingEnabled(True)
        table_layout.addWidget(self.table)

        layout.addWidget(table_card)

    def _fetch(self) -> None:
        access_token = (self.config.access_token or "").strip()
        if not access_token:
            self._logger.error("Basalam fetch blocked: access_token missing")
            dialogs.show_error(
                self,
                "Basalam Token Missing",
                "Set access_token in config.json before fetching.",
            )
            return

        vendor_id = "563284"
        if self._worker_thread and self._worker_thread.isRunning():
            self._logger.warning("Basalam fetch already running")
            return

        start_paid_at = self.start_paid_input.to_gregorian_str()
        end_paid_at = self.end_paid_input.to_gregorian_str()

        self._set_loading(True)
        self._logger.info(
            "Basalam fetch started vendor=%s start=%s end=%s",
            vendor_id,
            start_paid_at,
            end_paid_at,
        )

        self._worker_thread = QThread(self)
        self._worker = BasalamWorker(
            vendor_id=vendor_id,
            start_paid_at=start_paid_at,
            end_paid_at=end_paid_at,
            access_token=access_token,
            tabs=list(TABS_TO_FETCH),
            completed_status=TARGET_STATUS_FA,
            sleep_seconds=REQUEST_SLEEP_SECONDS,
        )
        self._worker.moveToThread(self._worker_thread)
        self._worker_thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.error.connect(self._on_worker_error)
        self._worker.finished.connect(self._worker_thread.quit)
        self._worker.error.connect(self._worker_thread.quit)
        self._worker_thread.finished.connect(self._worker.deleteLater)
        self._worker_thread.finished.connect(self._worker_thread.deleteLater)
        self._worker_thread.finished.connect(self._on_worker_thread_finished)
        self._worker_thread.start()

    def _export(self) -> None:
        if self._dataframe is None or self._dataframe.empty:
            self._logger.info("Basalam export skipped: no data")
            return

        from PySide6.QtWidgets import QFileDialog

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Basalam Orders",
            "basalam_orders.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
        )
        if not file_path:
            self._logger.info("Basalam export cancelled")
            return

        export_df, group_sizes = self._build_export_payload(self._dataframe)
        if file_path.lower().endswith(".csv"):
            export_df.to_csv(file_path, index=False)
        else:
            export_df.to_excel(file_path, index=False)
            self._apply_export_merges(file_path, export_df, group_sizes)
        self._logger.info(
            "Basalam export completed path=%s rows=%s",
            file_path,
            len(export_df),
        )

    def _on_progress(self, total_count: int) -> None:
        self.progress_label.setText(f"Loading... fetched {total_count} orders")

    def _on_worker_error(self, message: str) -> None:
        self._logger.error("Basalam worker error: %s", message)
        self._set_loading(False)
        dialogs.show_error(self, "Basalam Error", message)

    def _on_worker_finished(
        self, records: list, total_count: int, skipped_existing: int
    ) -> None:
        try:
            df = self._records_to_dataframe(records)
            self._dataframe = df
            self._update_table(df)
            self.export_button.setEnabled(df is not None and not df.empty)
            if total_count == 0:
                self.summary_label.setText("No data returned.")
            elif df.empty and skipped_existing:
                self.summary_label.setText(
                    f"{total_count} rows fetched, {skipped_existing} already processed."
                )
            else:
                if skipped_existing:
                    self.summary_label.setText(
                        f"{len(df)} new rows loaded, {skipped_existing} already processed."
                    )
                else:
                    self.summary_label.setText(f"{len(df)} rows loaded.")
            self._logger.info(
                "Basalam fetch finished fetched=%s new=%s skipped_existing=%s",
                total_count,
                len(df) if df is not None else 0,
                skipped_existing,
            )
        except Exception as exc:  # noqa: BLE001
            self._logger.exception("Basalam finished handler failed")
            dialogs.show_error(self, "Basalam Error", str(exc))
        finally:
            self._set_loading(False)

    def _on_worker_thread_finished(self) -> None:
        self._logger.info("Basalam worker thread finished cleanup")
        self._worker = None
        self._worker_thread = None

    def _set_loading(self, active: bool) -> None:
        if active:
            self.fetch_button.setEnabled(False)
            self.export_button.setEnabled(False)
            self.progress_bar.setRange(0, 0)
            self.progress_label.setText("Loading...")
            self.progress_bar.show()
            self.progress_label.show()
        else:
            self.fetch_button.setEnabled(True)
            self.progress_bar.hide()
            self.progress_label.hide()

    def _records_to_dataframe(self, records):
        import pandas as pd

        rows = self._extract_item_rows(records)
        columns = ["Recipient Name", "Product Name", "Color", "Quantity"]
        if not rows:
            return pd.DataFrame(columns=columns)
        return pd.DataFrame(rows, columns=columns)

    def _build_export_payload(self, df):
        import pandas as pd

        columns = [
            "Product Name",
            "Recipient Name",
            "Quantity",
            "Total Quantity",
        ]
        if df is None or df.empty:
            return pd.DataFrame(columns=columns), []

        if (
            "Product Name" not in df.columns
            or "Quantity" not in df.columns
            or "Recipient Name" not in df.columns
        ):
            return pd.DataFrame(columns=columns), []

        product_col = "Product Name"
        recipient_col = "Recipient Name"
        quantity_col = "Quantity"

        grouped_rows: dict[str, list[tuple[str, object, object]]] = {}
        product_order: list[str] = []

        totals: dict[str, int] = {}
        numeric_counts: dict[str, int] = {}

        for _, row in df.iterrows():
            product_value = row.get(product_col, "")
            if product_value is None or pd.isna(product_value):
                product_value = ""
            product = str(product_value).strip()

            recipient_value = row.get(recipient_col, "")
            if recipient_value is None or pd.isna(recipient_value):
                recipient_value = ""
            recipient = str(recipient_value).strip()

            quantity_value = row.get(quantity_col, "")
            quantity = self._coerce_quantity_value(quantity_value)

            if product not in grouped_rows:
                grouped_rows[product] = []
                product_order.append(product)
            grouped_rows[product].append((recipient, quantity, quantity_value))

            numeric_quantity = self._numeric_quantity(quantity_value)
            if numeric_quantity is not None:
                totals[product] = totals.get(product, 0) + numeric_quantity
                numeric_counts[product] = numeric_counts.get(product, 0) + 1

        export_rows: list[dict[str, object]] = []
        group_sizes: list[int] = []
        for product in product_order:
            quantities = grouped_rows.get(product, [])
            total_quantity = totals.get(product)
            has_numeric = numeric_counts.get(product, 0) > 0
            total_cell: object = (
                total_quantity
                if has_numeric and total_quantity is not None
                else ""
            )
            for idx, (recipient, quantity, _raw_value) in enumerate(quantities):
                export_rows.append(
                    {
                        "Product Name": product,
                        "Recipient Name": recipient,
                        "Quantity": quantity,
                        "Total Quantity": total_cell if idx == 0 else "",
                    }
                )
            group_sizes.append(len(quantities))

        return pd.DataFrame(export_rows, columns=columns), group_sizes

    @staticmethod
    def _apply_export_merges(
        file_path: str, df, group_sizes: list[int]
    ) -> None:
        if not group_sizes:
            return
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
        except ImportError:
            return

        if "Total Quantity" not in df.columns:
            return

        total_col = list(df.columns).index("Total Quantity") + 1
        start_row = 2

        wb = load_workbook(file_path)
        ws = wb.active
        for group_size in group_sizes:
            if group_size <= 0:
                continue
            end_row = start_row + group_size - 1
            if group_size > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=total_col,
                    end_row=end_row,
                    end_column=total_col,
                )
            cell = ws.cell(row=start_row, column=total_col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = end_row + 1

        wb.save(file_path)

    @staticmethod
    def _coerce_quantity_value(value):
        if value is None:
            return ""
        if isinstance(value, float) and math.isnan(value):
            return ""
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return ""
            if cleaned.isdigit():
                return int(cleaned)
            return cleaned
        return value

    @staticmethod
    def _numeric_quantity(value) -> int | None:
        if value is None:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return int(value)
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned.isdigit():
                return int(cleaned)
        return None

    def _extract_summary_rows(self, records: list[dict]) -> list[dict]:
        rows: list[dict] = []
        for record in records:
            if not isinstance(record, dict):
                continue
            customer_name = self._get_customer_name(record)
            items = self._get_items(record)
            if items:
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    product_name = self._get_product_name(item)
                    quantity = self._get_quantity(item)
                    if product_name is None and quantity is None:
                        continue
                    rows.append(
                        {
                            "Customer Name": customer_name or "",
                            "Product Name": product_name or "",
                            "Quantity": "" if quantity is None else quantity,
                        }
                    )
                continue

            product_name = self._get_product_name(record)
            quantity = self._get_quantity(record)
            if not (customer_name or product_name or quantity is not None):
                continue
            rows.append(
                {
                    "Customer Name": customer_name or "",
                    "Product Name": product_name or "",
                    "Quantity": "" if quantity is None else quantity,
                }
            )

        return rows

    def _extract_item_rows(self, records: list[dict]) -> list[dict]:
        rows: list[dict] = []
        for record in records:
            if not isinstance(record, dict):
                continue
            recipient_name = self._get_recipient_name(record)
            items = self._get_items(record)
            if not items:
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                product_name = self._get_product_name(item)
                color = self._get_item_color(item)
                quantity = self._get_quantity(item)
                rows.append(
                    {
                        "Recipient Name": recipient_name or "",
                        "Product Name": product_name or "",
                        "Color": color or "",
                        "Quantity": "" if quantity is None else quantity,
                    }
                )
        return rows

    def _get_items(self, record: dict) -> list[dict]:
        for key in (
            "items",
            "order_items",
            "orderItems",
            "line_items",
            "lines",
            "products",
        ):
            value = record.get(key)
            if isinstance(value, list):
                return value
        return []

    def _get_customer_name(self, record: dict) -> str | None:
        for key in ("customer", "buyer", "user", "client"):
            value = record.get(key)
            name = self._extract_name_from_value(value)
            if name:
                return name
        for key in (
            "customer_data_user_name",
            "customer_data_userName",
            "customerDataUserName",
            "customer_data_user",
            "customerDataUser",
        ):
            value = record.get(key)
            name = self._extract_name_from_value(value)
            if name:
                return name
        for path in (
            ("customer_data", "user", "name"),
            ("customer_data", "user_name"),
            ("customer_data", "userName"),
            ("customer_data", "user", "full_name"),
            ("customer_data", "user", "fullName"),
            ("customerData", "user", "name"),
            ("customerData", "user_name"),
            ("customerData", "userName"),
        ):
            value = self._get_nested_value(record, path)
            name = self._extract_name_from_value(value)
            if name:
                return name
        for key in (
            "customer_name",
            "buyer_name",
            "user_name",
            "full_name",
            "fullname",
        ):
            value = record.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    def _get_recipient_name(self, record: dict) -> str | None:
        for path in (
            ("customer_data", "recipient", "name"),
            ("customerData", "recipient", "name"),
        ):
            value = self._get_nested_value(record, path)
            if isinstance(value, str) and value.strip():
                return value.strip()
        for key in ("customer_data_recipient_name", "recipient_name"):
            value = record.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    def _extract_name_from_value(self, value) -> str | None:
        if isinstance(value, str) and value.strip():
            return value.strip()
        if isinstance(value, dict):
            for key in (
                "full_name",
                "fullname",
                "name",
                "title",
                "user_name",
                "userName",
            ):
                name = value.get(key)
                if isinstance(name, str) and name.strip():
                    return name.strip()
            first = value.get("first_name") or value.get("firstname")
            last = value.get("last_name") or value.get("lastname")
            if isinstance(first, str) and isinstance(last, str):
                full = f"{first.strip()} {last.strip()}".strip()
                return full if full else None
        return None

    @staticmethod
    def _get_nested_value(record: dict, path: tuple[str, ...]):
        current = record
        for key in path:
            if not isinstance(current, dict):
                return None
            if key not in current:
                return None
            current = current[key]
        return current

    def _get_product_name(self, record: dict) -> str | None:
        value = record.get("product")
        name = self._extract_name_from_value(value)
        if name:
            return name
        for key in (
            "product_name",
            "productTitle",
            "product_title",
            "title",
            "name",
        ):
            value = record.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    def _get_quantity(self, record: dict):
        for key in ("quantity", "qty", "count", "amount", "number"):
            value = record.get(key)
            if value is None:
                continue
            if isinstance(value, (int, float)):
                return int(value) if isinstance(value, bool) is False else value
            if isinstance(value, str):
                cleaned = value.strip()
                if cleaned.isdigit():
                    return int(cleaned)
                return cleaned
        return None

    def _get_item_color(self, item: dict) -> str | None:
        variation = item.get("variation")
        if not isinstance(variation, dict):
            return None
        properties = variation.get("properties")
        if not isinstance(properties, list):
            return None
        for prop in properties:
            if not isinstance(prop, dict):
                continue
            prop_info = prop.get("property")
            value_info = prop.get("value")
            title = (
                prop_info.get("title")
                if isinstance(prop_info, dict)
                else prop.get("title")
            )
            if isinstance(title, str) and title.strip() in {"رنگ", "Color"}:
                if isinstance(value_info, dict):
                    value_title = value_info.get("title")
                    if isinstance(value_title, str) and value_title.strip():
                        return value_title.strip()
                if isinstance(value_info, str) and value_info.strip():
                    return value_info.strip()
        return None

    @staticmethod
    def _pretty_column(name: str) -> str:
        cleaned = (
            str(name).replace("__", " ").replace(".", " ").replace("_", " ")
        )
        tokens = []
        for token in cleaned.split():
            lowered = token.lower()
            if lowered in {"id", "sku", "url"}:
                tokens.append(lowered.upper())
            else:
                tokens.append(token.capitalize())
        return " ".join(tokens)

    @staticmethod
    def _format_nested(value):
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _update_table(self, df) -> None:
        if df is None or df.empty:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.summary_label.setText("No data returned.")
            return

        max_rows = len(df)
        self.table.setColumnCount(len(df.columns))
        self.table.setRowCount(max_rows)
        self.table.setHorizontalHeaderLabels([str(c) for c in df.columns])

        for row_idx in range(max_rows):
            row = df.iloc[row_idx]
            for col_idx, value in enumerate(row):
                column_name = df.columns[col_idx]
                item = QTableWidgetItem(self._format_cell(value, column_name))
                if isinstance(value, (int, float)) and not self._is_nan(value):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row_idx, col_idx, item)

        self.table.resizeColumnsToContents()
        if len(df) > max_rows:
            self.summary_label.setText(f"Showing {max_rows} of {len(df)} rows.")
        else:
            self.summary_label.setText(f"{len(df)} rows loaded.")

    @staticmethod
    def _format_cell(value, column_name: object | None = None) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and math.isnan(value):
            return ""
        if column_name is not None and is_price_column(column_name):
            return format_amount(value)
        return str(value)

    @staticmethod
    def _is_nan(value: float) -> bool:
        return isinstance(value, float) and math.isnan(value)
