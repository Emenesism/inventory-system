from __future__ import annotations

from PySide6.QtCore import (
    QAbstractTableModel,
    QCoreApplication,
    QModelIndex,
    QObject,
    Qt,
    QThread,
    QTimer,
    Signal,
    Slot,
)
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from app.core.config import AppConfig
from app.models.errors import InventoryFileError
from app.services.action_log_service import ActionLogService
from app.services.inventory_service import InventoryService
from app.utils.excel import autofit_columns, ensure_sheet_rtl
from app.utils.numeric import format_amount
from app.utils.text import display_text


def _t(text: str) -> str:
    return QCoreApplication.translate("LowStockPage", text)


COL_PRODUCT = _t("نام کالا")
COL_QUANTITY = _t("تعداد")
COL_ALARM = _t("حد هشدار")
COL_NEEDED = _t("نیاز")
COL_AVG_BUY = _t("میانگین خرید")
COL_SOURCE = _t("منبع")


class _LowStockLoadWorker(QObject):
    loaded = Signal(object)
    finished = Signal()

    def __init__(
        self,
        inventory_service: InventoryService,
        threshold: int,
    ) -> None:
        super().__init__()
        self._inventory_service = inventory_service
        self._threshold = threshold

    @Slot()
    def run(self) -> None:
        try:
            rows = self._inventory_service.get_low_stock_rows(self._threshold)
        except InventoryFileError:
            rows = []
        self.loaded.emit(rows)
        self.finished.emit()


class _LowStockTableModel(QAbstractTableModel):
    def __init__(self, headers: list[str]) -> None:
        super().__init__()
        self._headers = headers
        self._rows: list[dict[str, object]] = []

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return len(self._headers)

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):  # noqa: ANN001
        if not index.isValid():
            return None
        row = index.row()
        col = index.column()
        if row < 0 or row >= len(self._rows):
            return None
        item = self._rows[row]
        if role == Qt.DisplayRole:
            if col == 0:
                return str(item.get("product", ""))
            if col == 1:
                return str(int(item.get("quantity", 0) or 0))
            if col == 2:
                return str(int(item.get("alarm", 0) or 0))
            if col == 3:
                return str(int(item.get("needed", 0) or 0))
            if col == 4:
                return format_amount(float(item.get("avg_buy", 0.0) or 0.0))
            if col == 5:
                return str(item.get("source", ""))
            return ""
        if role == Qt.TextAlignmentRole:
            if col in (0, 5):
                return Qt.AlignVCenter | Qt.AlignRight | Qt.AlignAbsolute
            return Qt.AlignCenter
        if role == Qt.BackgroundRole:
            qty = int(item.get("quantity", 0) or 0)
            alarm = int(item.get("alarm", 0) or 0)
            return self._severity_brush(qty, alarm)
        return None

    def headerData(
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.DisplayRole,
    ):  # noqa: ANN001, N802
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            if 0 <= section < len(self._headers):
                return self._headers[section]
            return None
        return str(section + 1)

    def set_rows(self, rows: list[dict[str, object]]) -> None:
        self.beginResetModel()
        self._rows = list(rows)
        self.endResetModel()

    @staticmethod
    def _severity_brush(qty: int, alarm: int) -> QBrush | None:
        if alarm <= 0:
            return None
        deficit = max(alarm - qty, 0)
        severity = min(deficit / alarm, 1.0)
        if severity <= 0:
            return None
        low = (255, 244, 228)
        high = (255, 153, 153)
        red = int(low[0] + (high[0] - low[0]) * severity)
        green = int(low[1] + (high[1] - low[1]) * severity)
        blue = int(low[2] + (high[2] - low[2]) * severity)
        return QBrush(QColor(red, green, blue))


class LowStockPage(QWidget):
    _PRODUCT_COL = 0
    _QUANTITY_COL = 1
    _ALARM_COL = 2
    _NEEDED_COL = 3
    _AVG_BUY_COL = 4
    _SOURCE_COL = 5
    _PRODUCT_MIN_WIDTH = 170
    _PRODUCT_MAX_WIDTH = 560
    _SOURCE_MIN_WIDTH = 96
    _SOURCE_MAX_WIDTH = 320
    _NUMERIC_MIN_WIDTH = 70
    _NUMERIC_MAX_WIDTH = 170

    def __init__(
        self,
        inventory_service: InventoryService,
        config: AppConfig,
        action_log_service: ActionLogService | None = None,
        current_admin_provider=None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.inventory_service = inventory_service
        self.config = config
        self.action_log_service = action_log_service
        self._current_admin_provider = current_admin_provider
        self._rows: list[dict[str, object]] = []
        self._load_thread: QThread | None = None
        self._load_worker: _LowStockLoadWorker | None = None
        self._pending_refresh = False
        self._controls_enabled = True
        self._fit_timer = QTimer(self)
        self._fit_timer.setSingleShot(True)
        self._fit_timer.setInterval(90)
        self._fit_timer.timeout.connect(self._fit_columns)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        header = QHBoxLayout()
        title = QLabel(self.tr("هشدار کمبود موجودی"))
        title.setStyleSheet("font-size: 20px; font-weight: 600;")
        header.addWidget(title)
        header.addStretch(1)

        self.export_button = QPushButton(self.tr("خروجی"))
        self.export_button.clicked.connect(self._export)
        header.addWidget(self.export_button)

        self.refresh_button = QPushButton(self.tr("بروزرسانی"))
        self.refresh_button.clicked.connect(self.refresh)
        header.addWidget(self.refresh_button)
        layout.addLayout(header)

        summary_card = QFrame()
        summary_card.setObjectName("Card")
        summary_layout = QHBoxLayout(summary_card)
        summary_layout.setContentsMargins(16, 16, 16, 16)
        summary_layout.setSpacing(24)

        self.items_label = QLabel(self.tr("کالاهای زیر حد هشدار: 0"))
        self.total_needed_label = QLabel(self.tr("مجموع نیاز: 0"))
        summary_layout.addWidget(self.items_label)
        summary_layout.addWidget(self.total_needed_label)
        summary_layout.addStretch(1)
        layout.addWidget(summary_card)

        table_card = QFrame()
        table_card.setObjectName("Card")
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(16, 16, 16, 16)

        headers = [
            self.tr("کالا"),
            self.tr("تعداد"),
            self.tr("هشدار"),
            self.tr("نیاز"),
            self.tr("میانگین خرید"),
            self.tr("منبع"),
        ]
        self._table_model = _LowStockTableModel(headers)
        self.table = QTableView()
        self.table.setModel(self._table_model)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setWordWrap(False)
        self.table.setTextElideMode(Qt.ElideRight)
        header_view = self.table.horizontalHeader()
        for col in range(self._table_model.columnCount()):
            header_view.setSectionResizeMode(col, QHeaderView.Interactive)
        header_view.setStretchLastSection(False)
        header_view.setMinimumSectionSize(64)
        self.table.verticalHeader().setDefaultSectionSize(32)
        if hasattr(self.table, "setUniformRowHeights"):
            self.table.setUniformRowHeights(True)
        table_layout.addWidget(self.table)
        layout.addWidget(table_card)

    def set_enabled_state(self, enabled: bool) -> None:
        self._controls_enabled = bool(enabled)
        self.table.setEnabled(enabled)
        self.refresh_button.setEnabled(enabled and self._load_thread is None)
        self.export_button.setEnabled(enabled and bool(self._rows))

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        self._defer_fit_columns()

    def showEvent(self, event) -> None:  # noqa: N802
        super().showEvent(event)
        self.request_layout_refresh()

    def refresh(self) -> None:
        if not self.inventory_service.is_loaded():
            self._rows = []
            self._table_model.set_rows([])
            self.items_label.setText(self.tr("کالاهای زیر حد هشدار: 0"))
            self.total_needed_label.setText(self.tr("مجموع نیاز: 0"))
            self.export_button.setEnabled(False)
            return
        if self._load_thread is not None and self._load_thread.isRunning():
            self._pending_refresh = True
            return
        self.refresh_button.setEnabled(False)
        worker = _LowStockLoadWorker(
            self.inventory_service,
            int(self.config.low_stock_threshold),
        )
        thread = QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.loaded.connect(self._on_rows_loaded)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(self._on_refresh_finished)
        thread.finished.connect(thread.deleteLater)
        self._load_worker = worker
        self._load_thread = thread
        thread.start()

    @Slot(object)
    def _on_rows_loaded(self, payload: object) -> None:
        rows_data = payload if isinstance(payload, list) else []
        rows: list[dict[str, object]] = []
        for row in rows_data:
            if not isinstance(row, dict):
                continue
            rows.append(
                {
                    "product": str(row.get("product_name", "")).strip(),
                    "quantity": int(row.get("quantity", 0) or 0),
                    "alarm": int(row.get("alarm", 0) or 0),
                    "needed": int(row.get("needed", 0) or 0),
                    "avg_buy": float(row.get("avg_buy_price", 0.0) or 0.0),
                    "source": display_text(row.get("source", ""), fallback=""),
                }
            )
        self._rows = rows
        self._table_model.set_rows(rows)
        self.items_label.setText(
            self.tr("کالاهای زیر حد هشدار: {count}").format(count=len(rows))
        )
        self.total_needed_label.setText(
            self.tr("مجموع نیاز: {count}").format(
                count=sum(item["needed"] for item in rows)
            )
        )
        self.export_button.setEnabled(self._controls_enabled and bool(rows))
        self._defer_fit_columns()

    def _on_refresh_finished(self) -> None:
        self._load_worker = None
        self._load_thread = None
        self.refresh_button.setEnabled(self._controls_enabled)
        self.export_button.setEnabled(
            self._controls_enabled and bool(self._rows)
        )
        if self._pending_refresh:
            self._pending_refresh = False
            self.refresh()

    def request_layout_refresh(self) -> None:
        self._defer_fit_columns()
        QTimer.singleShot(120, self._defer_fit_columns)

    def _export(self) -> None:
        if not self._rows:
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("خروجی لیست کمبود موجودی"),
            "low_stock.xlsx",
            self.tr("فایل‌های اکسل (*.xlsx);;فایل‌های CSV (*.csv)"),
        )
        if not file_path:
            return

        import pandas as pd

        df = pd.DataFrame(self._rows)
        df = df.rename(
            columns={
                "product": COL_PRODUCT,
                "quantity": COL_QUANTITY,
                "alarm": COL_ALARM,
                "needed": COL_NEEDED,
                "avg_buy": COL_AVG_BUY,
                "source": COL_SOURCE,
            }
        )
        if file_path.lower().endswith(".csv"):
            df.to_csv(file_path, index=False)
        else:
            df.to_excel(file_path, index=False)
            self._apply_export_colors(file_path, df)
            ensure_sheet_rtl(file_path)
            autofit_columns(file_path)
        if self.action_log_service:
            admin = (
                self._current_admin_provider()
                if self._current_admin_provider
                else None
            )
            self.action_log_service.log_action(
                "low_stock_export",
                self.tr("خروجی کمبود موجودی"),
                self.tr("تعداد ردیف‌ها: {count}\nمسیر: {path}").format(
                    count=len(df), path=file_path
                ),
                admin=admin,
            )

    def _apply_export_colors(self, file_path: str, df) -> None:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
        except ImportError:
            return

        max_needed = 0
        if COL_NEEDED not in df.columns:
            return
        for value in df[COL_NEEDED].tolist():
            try:
                max_needed = max(max_needed, int(value))
            except (TypeError, ValueError):
                continue
        if max_needed <= 0:
            return

        wb = load_workbook(file_path)
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        needed_col = list(df.columns).index(COL_NEEDED) + 1
        start_row = 2
        end_row = start_row + len(df) - 1

        for row_idx in range(start_row, end_row + 1):
            needed_value = ws.cell(row_idx, needed_col).value
            try:
                needed = int(needed_value)
            except (TypeError, ValueError):
                continue
            severity = min(max(needed / max_needed, 0.0), 1.0)
            green = int(235 - (140 * severity))
            red = int(255 - (15 * (1 - severity)))
            fill = PatternFill(
                start_color=f"{red:02X}{green:02X}66",
                end_color=f"{red:02X}{green:02X}66",
                fill_type="solid",
            )
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = fill

        wb.save(file_path)

    def _fit_columns(self) -> None:
        model = self.table.model()
        if model is None:
            return
        column_count = int(model.columnCount())
        if column_count <= self._PRODUCT_COL:
            return
        header = self.table.horizontalHeader()
        viewport_width = self.table.viewport().width()
        if viewport_width <= 0:
            return

        widths: list[int] = []
        min_widths: list[int] = []
        max_widths: list[int] = []
        for col in range(column_count):
            self.table.resizeColumnToContents(col)
            natural = int(header.sectionSize(col))
            min_width = self._column_min_width(col)
            max_width = self._column_max_width(col, viewport_width)
            widths.append(max(min_width, min(natural, max_width)))
            min_widths.append(min_width)
            max_widths.append(max_width)

        delta = int(viewport_width - sum(widths))
        if delta > 0:
            delta = self._grow_columns(widths, max_widths, delta)
            if delta > 0:
                # Always consume remaining space so table keeps full-width layout.
                widths[self._product_fill_col(widths)] += int(delta)
        elif delta < 0:
            self._shrink_columns(widths, min_widths, -delta)

        for col, width in enumerate(widths):
            header.resizeSection(col, int(max(40, width)))

    def _defer_fit_columns(self) -> None:
        self._fit_timer.start()

    def _column_min_width(self, col: int) -> int:
        if col == self._PRODUCT_COL:
            return self._PRODUCT_MIN_WIDTH
        if col == self._SOURCE_COL:
            return self._SOURCE_MIN_WIDTH
        if col == self._AVG_BUY_COL:
            return max(self._NUMERIC_MIN_WIDTH, 98)
        return self._NUMERIC_MIN_WIDTH

    def _column_max_width(self, col: int, viewport_width: int) -> int:
        if col == self._PRODUCT_COL:
            return max(
                260, min(self._PRODUCT_MAX_WIDTH, int(viewport_width * 0.52))
            )
        if col == self._SOURCE_COL:
            return max(
                self._SOURCE_MIN_WIDTH,
                min(self._SOURCE_MAX_WIDTH, int(viewport_width * 0.24)),
            )
        if col == self._AVG_BUY_COL:
            return max(
                120,
                min(self._NUMERIC_MAX_WIDTH + 40, int(viewport_width * 0.2)),
            )
        return max(
            self._NUMERIC_MIN_WIDTH,
            min(self._NUMERIC_MAX_WIDTH, int(viewport_width * 0.14)),
        )

    def _grow_columns(
        self, widths: list[int], max_widths: list[int], extra: int
    ) -> int:
        grow_order = [
            self._PRODUCT_COL,
            self._SOURCE_COL,
            self._AVG_BUY_COL,
            self._QUANTITY_COL,
            self._ALARM_COL,
            self._NEEDED_COL,
        ]
        for col in grow_order:
            if extra <= 0 or col >= len(widths):
                break
            capacity = max_widths[col] - widths[col]
            if capacity <= 0:
                continue
            added = min(capacity, extra)
            widths[col] += int(added)
            extra -= int(added)
        return extra

    def _shrink_columns(
        self, widths: list[int], min_widths: list[int], deficit: int
    ) -> None:
        shrink_order = [
            self._PRODUCT_COL,
            self._SOURCE_COL,
            self._AVG_BUY_COL,
            self._QUANTITY_COL,
            self._ALARM_COL,
            self._NEEDED_COL,
        ]
        while deficit > 0:
            progressed = False
            for col in shrink_order:
                if col >= len(widths):
                    continue
                available = widths[col] - min_widths[col]
                if available <= 0:
                    continue
                shrink_by = min(available, deficit)
                widths[col] -= int(shrink_by)
                deficit -= int(shrink_by)
                progressed = True
                if deficit <= 0:
                    break
            if not progressed:
                break

    def _product_fill_col(self, widths: list[int]) -> int:
        if self._PRODUCT_COL < len(widths):
            return self._PRODUCT_COL
        return 0
