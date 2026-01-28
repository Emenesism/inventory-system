from __future__ import annotations

from pathlib import Path

from app.utils.dates import to_jalali_datetime


def _ensure_sheet_direction(path: str | Path, right_to_left: bool) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    try:
        workbook = load_workbook(path)
    except Exception:  # noqa: BLE001
        return
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.rightToLeft = right_to_left
    try:
        workbook.save(path)
    except Exception:  # noqa: BLE001
        return


def ensure_sheet_ltr(path: str | Path) -> None:
    _ensure_sheet_direction(path, right_to_left=False)


def ensure_sheet_rtl(path: str | Path) -> None:
    _ensure_sheet_direction(path, right_to_left=True)


def apply_banded_rows(
    path: str | Path, header_row: int = 1, stripe_color: str = "F7F9FC"
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        return
    try:
        workbook = load_workbook(path)
    except Exception:  # noqa: BLE001
        return
    stripe_fill = PatternFill(
        start_color=stripe_color,
        end_color=stripe_color,
        fill_type="solid",
    )
    for worksheet in workbook.worksheets:
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        if max_row <= header_row or max_col < 1:
            continue
        start_row = header_row + 1
        for row_idx in range(start_row, max_row + 1):
            if (row_idx - start_row) % 2 == 0:
                for col_idx in range(1, max_col + 1):
                    worksheet.cell(
                        row=row_idx, column=col_idx
                    ).fill = stripe_fill
    try:
        workbook.save(path)
    except Exception:  # noqa: BLE001
        return


def autofit_columns(
    path: str | Path, min_width: int = 8, max_width: int = 50
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    try:
        workbook = load_workbook(path)
    except Exception:  # noqa: BLE001
        return
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                text_len = len(text.replace("\n", " "))
                if text_len > max_len:
                    max_len = text_len
            if max_len == 0:
                continue
            width = min(max(max_len + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = width
    try:
        workbook.save(path)
    except Exception:  # noqa: BLE001
        return


def _sanitize_sheet_title(value: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in value)
    cleaned = cleaned.strip()
    return cleaned[:31] if cleaned else "Invoice"


def _populate_invoice_sheet(ws, invoice, lines) -> None:
    try:
        from datetime import datetime
        from zoneinfo import ZoneInfo

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception:  # noqa: BLE001
        return

    ws.sheet_view.rightToLeft = True

    # Styles
    title_font = Font(name="Vazirmatn", size=18, bold=True)
    header_font = Font(name="Vazirmatn", size=11, bold=True)
    body_font = Font(name="Vazirmatn", size=11)
    label_font = Font(name="Vazirmatn", size=10, bold=True)
    header_fill = PatternFill(
        start_color="E8F3E1", end_color="E8F3E1", fill_type="solid"
    )
    stripe_fill = PatternFill(
        start_color="F7F9FC", end_color="F7F9FC", fill_type="solid"
    )
    thin = Side(border_style="thin", color="C7CED6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # Title
    title_text = (
        "فاکتور فروش" if invoice.invoice_type == "sales" else "فاکتور خرید"
    )
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    export_dt = datetime.now(ZoneInfo("Asia/Tehran"))
    export_date = to_jalali_datetime(
        export_dt.isoformat(timespec="seconds")
    ).split(" ")[0]
    invoice_date = to_jalali_datetime(invoice.created_at).split(" ")[0]

    # Header info
    ws["A3"].value = "تاریخ فاکتور:"
    ws["B3"].value = invoice_date
    ws["D3"].value = "تاریخ خروجی:"
    ws["E3"].value = export_date

    ws["A4"].value = "شماره فاکتور:"
    ws["B4"].value = str(invoice.invoice_id)
    ws["D4"].value = "نوع:"
    ws["E4"].value = "فروش" if invoice.invoice_type == "sales" else "خرید"

    for cell in ("A3", "D3", "A4", "D4"):
        ws[cell].font = label_font
    for cell in ("B3", "E3", "B4", "E4"):
        ws[cell].font = body_font
    for cell in ("A3", "B3", "D3", "E3", "A4", "B4", "D4", "E4"):
        ws[cell].alignment = Alignment(horizontal="right", vertical="center")

    # Table header
    header_row = 6
    headers = ["ردیف", "شرح کالا", "تعداد", "مبلغ واحد", "مبلغ کل"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Table rows
    total_qty = 0
    total_amount = 0.0
    for idx, line in enumerate(lines, start=1):
        row = header_row + idx
        qty = int(line.quantity)
        line_total = float(line.price) * qty
        total_qty += qty
        total_amount += line_total

        values = [
            idx,
            line.product_name,
            qty,
            float(line.price),
            line_total,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = body_font
            cell.border = border
            if col_idx in (1, 3):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            elif col_idx in (4, 5):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(
                    horizontal="right", vertical="center"
                )
        if idx % 2 == 0:
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).fill = stripe_fill

    # Totals row
    total_row = header_row + len(lines) + 1
    ws.cell(row=total_row, column=2, value="جمع کل").font = header_font
    ws.cell(row=total_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=total_row, column=3, value=total_qty).font = header_font
    ws.cell(row=total_row, column=3).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    total_cell = ws.cell(row=total_row, column=5, value=total_amount)
    total_cell.font = header_font
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.number_format = "#,##0"
    for col_idx in range(1, 6):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        cell.fill = PatternFill(
            start_color="EEF2FF", end_color="EEF2FF", fill_type="solid"
        )


def export_invoice_excel(file_path: str | Path, invoice, lines) -> None:
    try:
        from openpyxl import Workbook
    except Exception:  # noqa: BLE001
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    _populate_invoice_sheet(ws, invoice, lines)
    try:
        wb.save(file_path)
    except Exception:  # noqa: BLE001
        return
    autofit_columns(file_path, min_width=12, max_width=60)


def export_invoices_excel(file_path: str | Path, invoices_with_lines) -> None:
    try:
        from openpyxl import Workbook
    except Exception:  # noqa: BLE001
        return
    wb = Workbook()
    ws = wb.active
    used_titles: set[str] = set()
    for idx, (invoice, lines) in enumerate(invoices_with_lines):
        sheet = ws if idx == 0 else wb.create_sheet()
        base_title = _sanitize_sheet_title(str(invoice.invoice_id))
        title = base_title
        counter = 2
        while title in used_titles:
            suffix = f"_{counter}"
            title = (base_title[: 31 - len(suffix)] + suffix)[:31]
            counter += 1
        sheet.title = title
        used_titles.add(title)
        _populate_invoice_sheet(sheet, invoice, lines)
    try:
        wb.save(file_path)
    except Exception:  # noqa: BLE001
        return
    autofit_columns(file_path, min_width=12, max_width=60)
